from fastapi import APIRouter, Depends, HTTPException, Body, Query, Response
from fastapi.responses import StreamingResponse
from ipaddress import ip_network, ip_address
from pydantic import BaseModel
from datetime import datetime, timezone
from pathlib import Path
import csv
import io
import re

from sqlalchemy.orm import Session

from concurrent.futures import ThreadPoolExecutor
import threading
import socket

from ..database import SessionLocal

from .. import crud, schemas, models
from .. import utils
from ..utils import _clean_snmp_text, _normalize_ricoh_model, safe_int, detect_is_color, detect_is_color_from_model, get_db
from .. import ricoh_mib

from app.ricoh_http import get_ricoh_http_counters, get_ricoh_http_hostname
 
from ..snmp import get_snmp_value, get_snmp_values, get_snmp_walk
from .auth import require_tab
import time 

router = APIRouter(prefix="/printers", tags=["Printers"])


class CSVImportPayload(BaseModel):
    content: str
    filename: str | None = None


PRINTER_IMPORT_FIELDS = [
    "shared_name",
    "name",
    "model",
    "ip",
    "serial",
    "location",
    "is_color",
    "snmp_community",
]


def _csv_value(row: dict, *names: str) -> str:
    normalized = {str(k).strip().lower().replace(" ", "_"): v for k, v in row.items()}
    for name in names:
        value = normalized.get(name.strip().lower().replace(" ", "_"))
        if value is not None:
            return str(value).strip()
    return ""


def _parse_bool(value: str) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "si", "sí", "color", "y"}


def _decode_csv_content(content: str) -> list[dict]:
    text = (content or "").lstrip("\ufeff")
    try:
        dialect = csv.Sniffer().sniff(text[:2048], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="El archivo CSV no tiene encabezados")
    return list(reader)


def _downloads_dir() -> Path:
    downloads = Path.home() / "Downloads"
    downloads.mkdir(parents=True, exist_ok=True)
    return downloads


def _unique_download_path(filename: str) -> Path:
    path = _downloads_dir() / filename
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    for index in range(1, 1000):
        candidate = path.with_name(f"{stem}_{index}{suffix}")
        if not candidate.exists():
            return candidate
    return path.with_name(f"{stem}_{int(time.time())}{suffix}")


def _build_printers_csv(db: Session) -> str:
    output = io.StringIO()
    output.write("\ufeff")
    writer = csv.DictWriter(output, fieldnames=PRINTER_IMPORT_FIELDS)
    writer.writeheader()
    for printer in crud.get_printers(db):
        writer.writerow({
            "shared_name": printer.shared_name or "",
            "name": printer.name or "",
            "model": printer.model or "",
            "ip": printer.ip or "",
            "serial": printer.serial or "",
            "location": printer.location or "",
            "is_color": "true" if printer.is_color else "false",
            "snmp_community": printer.snmp_community or "public",
        })
    return output.getvalue()
STATUS_CACHE = {
    "timestamp": 0,
    "data": None
}

STATUS_CACHE_TTL = 65

MAX_WORKERS = 12
REFRESH_IN_PROGRESS = False
REFRESH_LOCK = threading.Lock()
COUNTERS_CACHE = {
    "timestamp": 0,
    "data": None
}
COUNTERS_CACHE_TTL = 60
COUNTERS_LOCK = threading.Lock()
COUNTERS_LAST_ROWS: dict[int, dict] = {}
COUNTERS_LAST_CHANGED_AT: dict[int, float] = {}
COUNTERS_LAST_REFRESH_TS = 0.0

HOSTNAME_SYNC_TTL = 300
HOSTNAME_SYNC_BATCH_SIZE = 2
HOSTNAME_LAST_SYNC: dict[int, float] = {}
HOSTNAME_SYNC_IN_PROGRESS: set[int] = set()
HOSTNAME_SYNC_LOCK = threading.Lock()

PRINTER_IDENTITY_OIDS = {
    "sys_descr": ricoh_mib.OID["sys_descr"],
    "sys_name": ricoh_mib.OID["sys_name"],
    "hr_descr_1": ricoh_mib.OID["hr_descr_1"],
    "hr_descr_2": ricoh_mib.OID["hr_descr_2"],
    "printer_name": ricoh_mib.OID["printer_name"],
    "serial": ricoh_mib.OID["printer_serial"],
    "toner_cyan": ricoh_mib.OID["marker_cyan"],
    "ricoh_sys_name": ricoh_mib.OID["ricoh_sys_name"],
    "ricoh_nic_name": ricoh_mib.OID["ricoh_nic_name"],
    "ricoh_nic_description": ricoh_mib.OID["ricoh_nic_description"],
}





def _looks_like_color_from_live_data(model, cyan_toner):
    if model:
        return detect_is_color_from_model(model)
    try:
        if cyan_toner is not None and int(cyan_toner) >= 0:
            return True
    except Exception:
        pass
    return None


def get_printer_identity(ip: str, community: str = "public", resolve_network_name: bool = False) -> dict:
    """Best-effort identity lookup for reconciling a reused printer IP."""
    oids = ricoh_mib.IDENTITY_OIDS
    values = get_snmp_values(ip, community, oids, timeout=2, retries=0)
    identity = ricoh_mib.parse_identity(values)
    name = identity.get("name")
    if resolve_network_name and not name:
        name = resolve_hostname_value(ip, community)

    return {
        "name": name,
        "model": identity.get("model"),
        "serial": identity.get("serial"),
        "is_color": identity.get("is_color"),
    }


def _build_reconcile_data(existing, incoming: dict, identity: dict) -> dict:
    data = {}
    for field in ("name", "model", "serial"):
        value = identity.get(field) or incoming.get(field)
        if value is not None and str(value).strip() and value != getattr(existing, field):
            data[field] = value

    if identity.get("is_color") is not None and identity.get("is_color") != existing.is_color:
        data["is_color"] = bool(identity.get("is_color"))
    elif incoming.get("model") and incoming.get("model") != "Desconocido" and "is_color" in incoming and incoming.get("is_color") != existing.is_color:
        data["is_color"] = bool(incoming.get("is_color"))

    if incoming.get("snmp_community") and incoming.get("snmp_community") != existing.snmp_community:
        data["snmp_community"] = incoming.get("snmp_community")

    if not existing.shared_name and incoming.get("shared_name"):
        data["shared_name"] = incoming.get("shared_name")
    if not existing.location and incoming.get("location"):
        data["location"] = incoming.get("location")

    return data


def _describe_printer_changes(existing, data: dict) -> str:
    labels = {
        "shared_name": "shared name",
        "name": "nombre",
        "model": "modelo",
        "serial": "serie",
        "ip": "IP",
        "location": "ubicacion",
        "is_color": "tipo",
        "snmp_community": "SNMP community",
    }
    parts = []
    for field, new_value in data.items():
        old_value = getattr(existing, field, None)
        if old_value == new_value:
            continue
        if field == "snmp_community":
            parts.append("SNMP community actualizada")
            continue
        if field == "is_color":
            old_text = "Color" if old_value else "B/N"
            new_text = "Color" if new_value else "B/N"
        else:
            old_text = str(old_value or "-")
            new_text = str(new_value or "-")
        parts.append(f"{labels.get(field, field)}: {old_text} -> {new_text}")
    return "; ".join(parts)


def _sync_hostname_for_printer(printer_id: int):
    db = SessionLocal()
    try:
        printer = crud.get_printer_by_id(db, printer_id)
        if not printer:
            return
        identity = get_printer_identity(printer.ip, printer.snmp_community, resolve_network_name=True)
        data = _build_reconcile_data(printer, {}, identity)
        if data:
            changes = _describe_printer_changes(printer, data)
            crud.update_printer(db, printer_id, data)
            if changes:
                crud.create_log(db, "printer", "updated", f'Identidad detectada por SNMP/HTTP en "{printer.shared_name or printer.name or printer.ip}": {changes}')
            clear_all_cache()
    except Exception:
        pass
    finally:
        db.close()
        with HOSTNAME_SYNC_LOCK:
            HOSTNAME_LAST_SYNC[printer_id] = time.time()
            HOSTNAME_SYNC_IN_PROGRESS.discard(printer_id)


def _schedule_hostname_sync(printer_id: int) -> bool:
    now = time.time()
    with HOSTNAME_SYNC_LOCK:
        last_sync = HOSTNAME_LAST_SYNC.get(printer_id, 0)
        if printer_id in HOSTNAME_SYNC_IN_PROGRESS:
            return False
        if (now - last_sync) < HOSTNAME_SYNC_TTL:
            return False
        HOSTNAME_SYNC_IN_PROGRESS.add(printer_id)

    threading.Thread(target=_sync_hostname_for_printer, args=(printer_id,), daemon=True).start()
    return True





def decode_printer_error_state(raw_value):

    if raw_value is None:

        return []

    try:

        octets = list(raw_value.asNumbers())

    except:

        try:

            octets = list(bytes(raw_value))

        except:

            return []

    bit_names = [

        "Low Paper",             # 0

        "No Paper",              # 1

        "Low Toner",             # 2

        "No Toner",              # 3

        "Door Open",             # 4

        "Paper Jam",             # 5

        "Offline",               # 6

        "Service Requested",     # 7

        "Input Tray Missing",    # 8

        "Output Tray Missing",   # 9

        "Marker Supply Missing", # 10

        "Output Near Full",      # 11

        "Output Full",           # 12

        "Input Tray Empty",      # 13

        "Maintenance Required"   # 14

    ]

    errors = []

    for byte_index, byte_value in enumerate(octets):

        for bit_index in range(8):

            global_bit_index = byte_index * 8 + bit_index

            if global_bit_index >= len(bit_names):

                continue

            mask = 0x80 >> bit_index

            if byte_value & mask:

                errors.append(bit_names[global_bit_index])

    return errors





def build_fast_printer_status(printer):

    model_hint = " ".join(
        str(getattr(printer, attr, "") or "")
        for attr in ("shared_name", "name", "model")
    )
    is_color = False if ricoh_mib.model_says_mono(model_hint) else detect_is_color(printer)

    oids = list(ricoh_mib.STATUS_OIDS)
    if is_color:
        oids += ricoh_mib.COLOR_STATUS_OIDS

    try:
        values = get_snmp_values(printer.ip, printer.snmp_community, oids)
    except Exception:
        values = None

    def _v(oid):
        if not values:
            return None
        return values.get(oid)

    # Use printer.name from DB (kept current by _sync_hostname_for_printer via
    # resolve_hostname_value: HTTP > DNS > SNMP). Reading sysName directly from
    # SNMP is unreliable on Ricoh printers — it often returns the model name.
    live_name = printer.name or printer.model

    toner_black = ricoh_mib.choose_toner(values or {}, "marker_black", "ricoh_toner_black")
    status_code = safe_int(_v(ricoh_mib.OID["printer_state"]))
    error_state_raw = _v(ricoh_mib.OID["printer_error_state"])
    error_list = decode_printer_error_state(error_state_raw)
    alert_message = ricoh_mib.parse_status_alert(values or {})

    snmp_alive = any(value is not None for value in (values or {}).values())
    status = "online" if snmp_alive else "offline"
    printer_state = "unknown"
    if status_code == 3:
        printer_state = "idle"
    elif status_code == 4:
        printer_state = "printing"
    elif status_code == 5:
        printer_state = "warmup"
    elif status_code == 6:
        printer_state = "stopped"

    if status == "offline":
        error_message = "Sin respuesta SNMP"
    elif alert_message:
        error_message = alert_message
    elif error_list:
        error_message = ", ".join(error_list)
    elif toner_black is not None and toner_black < 10:
        error_message = "Tóner crítico"
    elif toner_black is not None and toner_black < 30:
        error_message = "Low Toner"
    else:
        error_message = "Sin error"

    printer_data = {
        "id": printer.id,
        "shared_name": getattr(printer, "shared_name", None),
        "name": live_name,
        "model": printer.model,
        "serial": printer.serial,
        "ip": printer.ip,
        "location": printer.location,
        "is_color": is_color,
        "status": status,
        "printer_state": printer_state,
        "error_message": error_message,
        "toner_black": toner_black
    }

    if is_color:
        toner_cyan = ricoh_mib.choose_toner(values or {}, "marker_cyan", "ricoh_toner_cyan")
        toner_magenta = ricoh_mib.choose_toner(values or {}, "marker_magenta", "ricoh_toner_magenta")
        toner_yellow = ricoh_mib.choose_toner(values or {}, "marker_yellow", "ricoh_toner_yellow")
        printer_data["toner_cyan"] = toner_cyan
        printer_data["toner_magenta"] = toner_magenta
        printer_data["toner_yellow"] = toner_yellow

    return printer_data


def resolve_hostname_value(ip: str, community: str = "public") -> str | None:
    hostname = None

    try:
        snmp_name = get_snmp_value(
            ip,
            community,
            "1.3.6.1.2.1.1.5.0",
            timeout=2,
            retries=0,
        )
        if snmp_name is not None:
            hostname = str(snmp_name).strip()
    except Exception:
        hostname = None

    if not hostname:
        try:
            hostname = socket.gethostbyaddr(ip)[0]
        except Exception:
            hostname = None

    if not hostname:
        try:
            hostname = get_ricoh_http_hostname(ip)
        except Exception:
            hostname = None

    if hostname and "." in hostname:
        hostname = hostname.split(".")[0]

    return hostname


@router.post("/", response_model=schemas.PrinterResponse, dependencies=[Depends(require_tab("printers"))])
def create_printer(printer: schemas.PrinterCreate, db: Session = Depends(get_db)):
    existing = crud.get_printer_by_ip(db, printer.ip)
    if existing:
        incoming = printer.dict()
        identity = get_printer_identity(printer.ip, printer.snmp_community)
        data = _build_reconcile_data(existing, incoming, identity)
        if data:
            updated = crud.update_printer(db, existing.id, data)
            clear_all_cache()
            changed = ", ".join(sorted(data.keys()))
            crud.create_log(db, "printer", "updated", f'IP reutilizada: "{updated.shared_name or updated.name}" ({updated.ip}) reconciliada ({changed})')
            return updated
        return existing

    printer_data = printer.dict()
    identity = get_printer_identity(printer.ip, printer.snmp_community)
    for field in ("name", "model", "serial"):
        if identity.get(field):
            printer_data[field] = identity[field]
    if identity.get("is_color") is True:
        printer_data["is_color"] = True

    created = crud.create_printer(db, schemas.PrinterCreate(**printer_data))
    clear_all_cache()
    crud.create_log(db, "printer", "created", f'Impresora "{created.shared_name or created.name}" ({created.ip}) agregada')
    return created


@router.get("/", response_model=list[schemas.PrinterResponse], dependencies=[Depends(require_tab("dashboard", "printers", "tonerControl", "counters"))])
def list_printers(db: Session = Depends(get_db)):
    printers = crud.get_printers(db)
    scheduled = 0
    for printer in printers:
        if scheduled >= HOSTNAME_SYNC_BATCH_SIZE:
            break
        if _schedule_hostname_sync(printer.id):
            scheduled += 1
    return printers


@router.get("/export/csv", dependencies=[Depends(require_tab("printers"))])
def export_printers_csv(db: Session = Depends(get_db)):
    output = io.StringIO(_build_printers_csv(db))
    return StreamingResponse(
        output,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=printers.csv"},
    )


@router.get("/export/csv/downloads", dependencies=[Depends(require_tab("printers"))])
def export_printers_csv_to_downloads(db: Session = Depends(get_db)):
    path = _unique_download_path("printers.csv")
    path.write_text(_build_printers_csv(db), encoding="utf-8")
    return {"ok": True, "path": str(path)}


@router.post("/import/csv", dependencies=[Depends(require_tab("printers"))])
def import_printers_csv(payload: CSVImportPayload, db: Session = Depends(get_db)):
    created = 0
    updated = 0
    skipped = 0
    errors = []
    rows = _decode_csv_content(payload.content)

    for index, row in enumerate(rows, start=2):
        try:
            ip = _csv_value(row, "ip", "static_ip")
            model = _csv_value(row, "model", "modelo")
            if not ip:
                skipped += 1
                errors.append({"row": index, "error": "IP requerida"})
                continue
            if not model:
                model = "Desconocido"

            existing = crud.get_printer_by_ip(db, ip)

            data = {
                "shared_name": _csv_value(row, "shared_name", "shared name", "nombre_compartido") or None,
                "name": _csv_value(row, "name", "nombre") or None,
                "model": model,
                "ip": ip,
                "serial": _csv_value(row, "serial", "serie") or None,
                "location": _csv_value(row, "location", "ubicacion", "ubicación") or None,
                "is_color": _parse_bool(_csv_value(row, "is_color", "color")),
                "snmp_community": _csv_value(row, "snmp_community", "community") or "public",
            }

            if existing:
                duplicate = crud.get_printer_by_ip(db, ip)
                if duplicate and duplicate.id != existing.id:
                    skipped += 1
                    errors.append({"row": index, "error": f"IP duplicada: {ip}"})
                    continue
                crud.update_printer(db, existing.id, data)
                updated += 1
            else:
                crud.create_printer(db, schemas.PrinterCreate(**data))
                created += 1
        except Exception as exc:
            skipped += 1
            errors.append({"row": index, "error": str(exc)[:200]})

    clear_all_cache()
    crud.create_log(db, "printer", "imported", f"Importacion CSV de impresoras: {created} creadas, {updated} actualizadas, {skipped} omitidas")
    return {"created": created, "updated": updated, "skipped": skipped, "errors": errors}


@router.put("/{printer_id}", response_model=schemas.PrinterResponse, dependencies=[Depends(require_tab("printers"))])
def update_printer(printer_id: int, printer: schemas.PrinterUpdate, db: Session = Depends(get_db)):
    existing = crud.get_printer_by_id(db, printer_id)
    if not existing:
        raise HTTPException(status_code=404, detail="Printer not found")

    data = printer.dict(exclude_unset=True)
    if "ip" in data and data["ip"] != existing.ip:
        duplicate = crud.get_printer_by_ip(db, data["ip"])
        if duplicate:
            raise HTTPException(status_code=400, detail="Ya existe una impresora con esa IP")

    target_ip = data.get("ip", existing.ip)
    target_community = data.get("snmp_community", existing.snmp_community)
    identity = get_printer_identity(target_ip, target_community)
    for field in ("name", "model", "serial"):
        if identity.get(field):
            data[field] = identity[field]
    if identity.get("is_color") is not None:
        data["is_color"] = bool(identity.get("is_color"))

    changes = _describe_printer_changes(existing, data)
    updated = crud.update_printer(db, printer_id, data)
    clear_all_cache()
    detail = f": {changes}" if changes else ""
    crud.create_log(db, "printer", "updated", f'Impresora "{updated.shared_name or updated.name}" ({updated.ip}) actualizada{detail}')
    return updated


@router.delete("/{printer_id}", dependencies=[Depends(require_tab("printers"))])
def delete_printer(printer_id: int, db: Session = Depends(get_db)):
    printer = crud.get_printer_by_id(db, printer_id)
    if not printer:
        raise HTTPException(status_code=404, detail="Printer not found")

    label = f'Impresora "{printer.shared_name or printer.name}" ({printer.ip}) eliminada'
    crud.delete_printer(db, printer_id)
    clear_all_cache()
    crud.create_log(db, "printer", "deleted", label)
    return {"message": "Printer deleted"}


@router.post("/{printer_id}/reconcile", response_model=schemas.PrinterResponse, dependencies=[Depends(require_tab("printers"))])
def reconcile_printer_identity(printer_id: int, db: Session = Depends(get_db)):
    printer = crud.get_printer_by_id(db, printer_id)
    if not printer:
        raise HTTPException(status_code=404, detail="Printer not found")

    identity = get_printer_identity(printer.ip, printer.snmp_community)
    data = _build_reconcile_data(printer, {}, identity)
    if not data:
        return printer

    updated = crud.update_printer(db, printer_id, data)
    clear_all_cache()
    changed = _describe_printer_changes(printer, data) or ", ".join(sorted(data.keys()))
    crud.create_log(db, "printer", "updated", f'Identidad reconciliada: "{updated.shared_name or updated.name}" ({updated.ip}) ({changed})')
    return updated


@router.get("/status", dependencies=[Depends(require_tab("dashboard"))])
def get_printers_status(db: Session = Depends(get_db)):
    now = time.time()

    if (
        STATUS_CACHE["data"] is not None and
        (now - STATUS_CACHE["timestamp"]) < STATUS_CACHE_TTL
    ):
        return STATUS_CACHE["data"]

    printers = crud.get_printers(db)

    scheduled = 0
    for printer in printers:
        if scheduled >= HOSTNAME_SYNC_BATCH_SIZE:
            break
        if _schedule_hostname_sync(printer.id):
            scheduled += 1

    # If there is no cached data yet, return a lightweight placeholder and refresh in background
    if STATUS_CACHE["data"] is None:

        placeholder = []

        for p in printers:

            placeholder.append({

                "id": p.id,

                "shared_name": getattr(p, "shared_name", None),

                "name": p.name,

                "model": p.model,

                "serial": p.serial,

                "ip": p.ip,

                "location": p.location,

                "is_color": detect_is_color(p),

                "status": "unknown",

                "printer_state": "unknown",

                "error_message": "Cargando...",

                "toner_black": None

            })

        STATUS_CACHE["timestamp"] = now

        STATUS_CACHE["data"] = placeholder

        def refresh_cache():

            try:

                with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:

                    results = list(executor.map(build_fast_printer_status, printers))

                STATUS_CACHE["timestamp"] = time.time()

                STATUS_CACHE["data"] = results

            finally:

                with REFRESH_LOCK:

                    global REFRESH_IN_PROGRESS

                    REFRESH_IN_PROGRESS = False

        with REFRESH_LOCK:

            global REFRESH_IN_PROGRESS

            if not REFRESH_IN_PROGRESS:

                REFRESH_IN_PROGRESS = True

                threading.Thread(target=refresh_cache, daemon=True).start()

        return placeholder

    # otherwise perform synchronous refresh

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:

        results = list(executor.map(build_fast_printer_status, printers))

    STATUS_CACHE["timestamp"] = now

    STATUS_CACHE["data"] = results

    return results


def build_printer_counters(printer):
    # basic counters (batch SNMP)
    oids = ricoh_mib.COUNTER_OIDS
    model_hint = " ".join(
        str(getattr(printer, attr, "") or "")
        for attr in ("shared_name", "name", "model")
    )
    model_says_mono = ricoh_mib.model_says_mono(model_hint)
    try:
        vals = get_snmp_values(printer.ip, printer.snmp_community, oids)
    except Exception:
        vals = None

    def _v(oid):
        if not vals:
            return None
        return vals.get(oid)

    reading, is_color = ricoh_mib.normalize_counter_reading(
        vals or {},
        model_text=model_hint,
        db_is_color=None if model_says_mono else bool(getattr(printer, "is_color", False)),
    )
    total_pages = reading.total_pages
    bw_pages = reading.bw_pages
    color_pages = reading.color_pages

    snmp_present = any(_v(oid) is not None for oid in oids)

    # additional counters: copies and prints (may be available via HTTP on Ricoh devices)
    copy_bw = None
    copy_color = None
    print_bw = None
    print_color = None

    http_used = False
    http_data = get_ricoh_http_counters(printer.ip)
    if http_data:
        if http_data.get("copy_bw") is not None:
            copy_bw = safe_int(http_data["copy_bw"])
            http_used = True
        if is_color and http_data.get("copy_color") is not None:
            copy_color = safe_int(http_data["copy_color"])
            http_used = True
        if http_data.get("print_bw") is not None:
            print_bw = safe_int(http_data["print_bw"])
            http_used = True
        if is_color and http_data.get("print_color") is not None:
            print_color = safe_int(http_data["print_color"])
            http_used = True
        if http_data.get("bw_pages") is not None and (copy_bw is not None or print_bw is not None):
            bw_pages = safe_int(http_data["bw_pages"])
            http_used = True
        if is_color and http_data.get("color_pages") is not None and (copy_color is not None or print_color is not None):
            color_pages = safe_int(http_data["color_pages"])
            http_used = True
        if http_data.get("total_pages") is not None and (
            copy_bw is not None or print_bw is not None or copy_color is not None or print_color is not None
        ):
            total_pages = safe_int(http_data["total_pages"])
            http_used = True

    # if copy/print counters are available prefer summing them for BW and Color
    if copy_bw is not None or print_bw is not None:
        bw_pages = (copy_bw or 0) + (print_bw or 0)
    if is_color and (copy_color is not None or print_color is not None):
        color_pages = (copy_color or 0) + (print_color or 0)

    # Some Ricoh color models expose total and B/W counters but not the private
    # color counter. Never do this for models RMAdmin/our hints identify as mono.
    if is_color and not model_says_mono and color_pages is None and total_pages is not None and bw_pages is not None:
        derived_color = total_pages - bw_pages
        if derived_color >= 0:
            color_pages = derived_color

    # recompute total if possible
    if is_color:
        if (bw_pages is not None) and (color_pages is not None):
            total_pages = bw_pages + color_pages
    else:
        if bw_pages is None and total_pages is not None:
            bw_pages = total_pages
        elif bw_pages is not None:
            total_pages = bw_pages
        color_pages = None
        copy_color = None
        print_color = None

    counter_status = "online" if (total_pages is not None or bw_pages is not None or (is_color and color_pages is not None)) else "offline"
    if snmp_present and http_used:
        counter_source = "mixed"
    elif http_used:
        counter_source = "http"
    elif snmp_present:
        counter_source = "snmp"
    else:
        counter_source = "fallback"

    row = {
        "id": printer.id,
        "name": getattr(printer, "shared_name", None) or printer.name or printer.model,
        "model": printer.model,
        "serial": printer.serial,
        "ip": printer.ip,
        "is_color": is_color,
        "counter_status": counter_status,
        "counter_source": counter_source,
        "total_pages": total_pages,
        "copy_bw": copy_bw,
        "copy_color": copy_color,
        "print_bw": print_bw,
        "print_color": print_color,
        "bw_pages": bw_pages,
        "color_pages": color_pages,
    }
    row["is_complete"] = _counter_is_complete(row)
    return row


@router.get("/counters", dependencies=[Depends(require_tab("counters"))])
def get_printers_counters(
    status: str = Query(default="all", pattern="^(all|online|offline)$"),
    printer_type: str = Query(default="all", pattern="^(all|mono|color)$"),
    incomplete: str = Query(default="all", pattern="^(all|true|false)$"),
    sort_by: str = Query(default="name"),
    sort_dir: str = Query(default="asc", pattern="^(asc|desc)$"),
    force: bool = Query(default=False),
    db: Session = Depends(get_db),
):
    if force:
        rows, _ = _refresh_counters_cache(db)
    else:
        rows = _get_cached_counter_rows()
        if not rows:
            rows = _get_latest_counter_snapshot_rows(db)
        if not rows:
            rows, _ = _refresh_counters_cache(db)
    return _apply_counter_filters_and_sort(
        rows,
        status=status,
        printer_type=printer_type,
        incomplete=incomplete,
        sort_by=sort_by,
        sort_dir=sort_dir,
    )


@router.get("/counters/export", dependencies=[Depends(require_tab("counters"))])
def export_printers_counters_csv(
    status: str = Query(default="all", pattern="^(all|online|offline)$"),
    printer_type: str = Query(default="all", pattern="^(all|mono|color)$"),
    incomplete: str = Query(default="all", pattern="^(all|true|false)$"),
    sort_by: str = Query(default="name"),
    sort_dir: str = Query(default="asc", pattern="^(asc|desc)$"),
    db: Session = Depends(get_db),
):
    rows = _get_cached_counter_rows()
    if not rows:
        rows = _get_latest_counter_snapshot_rows(db)
    if not rows:
        rows, _ = _refresh_counters_cache(db)
    rows = _apply_counter_filters_and_sort(
        rows,
        status=status,
        printer_type=printer_type,
        incomplete=incomplete,
        sort_by=sort_by,
        sort_dir=sort_dir,
    )

    output = io.StringIO()
    writer = csv.writer(output)
    headers = [
        "id", "name", "model", "serial", "ip", "is_color", "counter_status", "counter_source",
        "total_pages", "copy_bw", "copy_color", "print_bw", "print_color", "bw_pages", "color_pages", "is_complete"
    ]
    writer.writerow(headers)
    for row in rows:
        writer.writerow([row.get(h) for h in headers])

    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=printers_counters.csv"},
    )


@router.get("/counters/changes", dependencies=[Depends(require_tab("counters"))])
def get_printers_counters_changes(since: float = Query(default=0), db: Session = Depends(get_db)):
    rows, refreshed_at = _refresh_counters_cache(db)
    with COUNTERS_LOCK:
        changed_ids = [pid for pid, changed_at in COUNTERS_LAST_CHANGED_AT.items() if changed_at > since]
    changed = [row for row in rows if row.get("id") in changed_ids]
    return {
        "server_ts": refreshed_at,
        "changed": changed,
        "full_refresh_required": since <= 0,
    }


@router.get("/{printer_id}/counters/debug", dependencies=[Depends(require_tab("counters"))])
def get_printer_counters_debug(printer_id: int, db: Session = Depends(get_db)):
    printer = crud.get_printer_by_id(db, printer_id)
    if not printer:
        raise HTTPException(status_code=404, detail="Printer not found")

    is_color = detect_is_color(printer)
    oids = ricoh_mib.COUNTER_OIDS
    snmp_raw: dict[str, str | None] = {}
    try:
        vals = get_snmp_values(printer.ip, printer.snmp_community, oids)
    except Exception:
        vals = None
    for oid in oids:
        raw = vals.get(oid) if vals else None
        snmp_raw[oid] = None if raw is None else str(raw)

    http_raw = None
    try:
        http_raw = get_ricoh_http_counters(printer.ip)
    except Exception:
        http_raw = None

    final_applied = build_printer_counters(printer)
    return {
        "printer": {
            "id": printer.id,
            "name": printer.name,
            "ip": printer.ip,
            "serial": printer.serial,
            "is_color": is_color,
        },
        "snmp_raw": snmp_raw,
        "snmp_map": {
            "printer_total": ricoh_mib.OID["printer_total"],
            "ricoh_bw_pages": ricoh_mib.OID["ricoh_bw_pages"],
            "ricoh_color_pages": ricoh_mib.OID["ricoh_color_pages"],
        },
        "http_raw": http_raw,
        "final_applied": final_applied,
    }


@router.get("/{printer_id}/counters/history", dependencies=[Depends(require_tab("counters", "tonerControl"))])
def get_printer_counters_history(
    printer_id: int,
    granularity: str = Query(default="daily", pattern="^(daily|weekly|monthly)$"),
    limit: int = Query(default=90, ge=1, le=500),
    db: Session = Depends(get_db),
):
    printer = crud.get_printer_by_id(db, printer_id)
    if not printer:
        raise HTTPException(status_code=404, detail="Printer not found")

    rows = db.query(models.PrinterCounterSnapshot).filter(
        models.PrinterCounterSnapshot.printer_id == printer_id,
        models.PrinterCounterSnapshot.granularity == granularity,
    ).order_by(models.PrinterCounterSnapshot.captured_at.desc()).limit(limit).all()

    return {
        "printer_id": printer_id,
        "granularity": granularity,
        "items": [
            {
                "captured_at": r.captured_at.isoformat() if r.captured_at else None,
                "period_bucket": r.period_bucket,
                "total_pages": r.total_pages,
                "bw_pages": r.bw_pages,
                "color_pages": r.color_pages,
                "copy_bw": r.copy_bw,
                "copy_color": r.copy_color,
                "print_bw": r.print_bw,
                "print_color": r.print_color,
                "source": r.source,
                "is_complete": r.is_complete,
            }
            for r in rows
        ],
    }


def clear_status_cache():

    STATUS_CACHE["timestamp"]=0

    STATUS_CACHE["data"]= None


def clear_counters_cache():

    COUNTERS_CACHE["timestamp"]=0

    COUNTERS_CACHE["data"]=None

    with COUNTERS_LOCK:
        global COUNTERS_LAST_REFRESH_TS
        COUNTERS_LAST_ROWS.clear()
        COUNTERS_LAST_CHANGED_AT.clear()
        COUNTERS_LAST_REFRESH_TS = 0.0


def clear_all_cache():
    clear_status_cache()
    clear_counters_cache()


@router.get("/resolve-hostname", dependencies=[Depends(require_tab("printers"))])
def resolve_printer_hostname(ip: str, community: str = "public"):
    identity = get_printer_identity(ip, community)
    return {
        "hostname": identity.get("name"),
        "name": identity.get("name"),
        "model": identity.get("model"),
        "serial": identity.get("serial"),
        "is_color": bool(identity.get("is_color")),
    }


def _counter_is_complete(row: dict) -> bool:
    if not row:
        return False
    is_color = bool(row.get("is_color"))
    if row.get("bw_pages") is None or row.get("total_pages") is None:
        return False
    if is_color and row.get("color_pages") is None:
        return False
    return True


def _snapshot_bucket(captured_at: datetime, granularity: str) -> str:
    if granularity == "daily":
        return captured_at.strftime("%Y-%m-%d")
    if granularity == "weekly":
        iso = captured_at.isocalendar()
        return f"{iso.year}-{iso.week:02d}"
    if granularity == "monthly":
        return captured_at.strftime("%Y-%m")
    raise ValueError("Granularity not supported")


def _upsert_counter_snapshots(db: Session, counters: list[dict], captured_at: datetime):
    for row in counters:
        printer_id = row.get("id")
        if printer_id is None:
            continue
        for granularity in ("daily", "weekly", "monthly"):
            bucket = _snapshot_bucket(captured_at, granularity)
            snap = db.query(models.PrinterCounterSnapshot).filter(
                models.PrinterCounterSnapshot.printer_id == printer_id,
                models.PrinterCounterSnapshot.granularity == granularity,
                models.PrinterCounterSnapshot.period_bucket == bucket,
            ).first()
            if not snap:
                snap = models.PrinterCounterSnapshot(
                    printer_id=printer_id,
                    granularity=granularity,
                    period_bucket=bucket,
                )
                db.add(snap)
            snap.captured_at = captured_at
            snap.total_pages = row.get("total_pages")
            snap.bw_pages = row.get("bw_pages")
            snap.color_pages = row.get("color_pages")
            snap.copy_bw = row.get("copy_bw")
            snap.copy_color = row.get("copy_color")
            snap.print_bw = row.get("print_bw")
            snap.print_color = row.get("print_color")
            snap.source = row.get("counter_source")
            snap.is_complete = bool(row.get("is_complete"))
    db.commit()


def _counter_signature(row: dict) -> tuple:
    return (
        row.get("total_pages"),
        row.get("bw_pages"),
        row.get("color_pages"),
        row.get("copy_bw"),
        row.get("copy_color"),
        row.get("print_bw"),
        row.get("print_color"),
        row.get("counter_status"),
        row.get("is_complete"),
    )


def _apply_counter_filters_and_sort(
    rows: list[dict],
    status: str = "all",
    printer_type: str = "all",
    incomplete: str = "all",
    sort_by: str = "name",
    sort_dir: str = "asc",
) -> list[dict]:
    filtered: list[dict] = []
    for row in rows or []:
        row_status = (row.get("counter_status") or "offline").lower()
        row_type = "color" if row.get("is_color") else "mono"
        row_incomplete = not bool(row.get("is_complete"))

        if status in {"online", "offline"} and row_status != status:
            continue
        if printer_type in {"mono", "color"} and row_type != printer_type:
            continue
        if incomplete == "true" and not row_incomplete:
            continue
        if incomplete == "false" and row_incomplete:
            continue
        filtered.append(row)

    reverse = (sort_dir or "asc").lower() == "desc"

    def sort_key(item: dict):
        val = item.get(sort_by)
        if isinstance(val, str):
            return val.lower()
        if val is None:
            return -1 if reverse else 10**18
        return val

    try:
        filtered.sort(key=sort_key, reverse=reverse)
    except Exception:
        pass
    return filtered


def _get_cached_counter_rows() -> list[dict]:
    rows = COUNTERS_CACHE.get("data")
    if rows:
        return rows
    with COUNTERS_LOCK:
        if COUNTERS_LAST_ROWS:
            return list(COUNTERS_LAST_ROWS.values())
    return []


def _get_latest_counter_snapshot_rows(db: Session) -> list[dict]:
    printers = crud.get_printers(db)
    snapshots = (
        db.query(models.PrinterCounterSnapshot)
        .filter(models.PrinterCounterSnapshot.granularity == "daily")
        .order_by(models.PrinterCounterSnapshot.captured_at.desc())
        .all()
    )
    latest_by_printer: dict[int, models.PrinterCounterSnapshot] = {}
    for snapshot in snapshots:
        if snapshot.printer_id not in latest_by_printer:
            latest_by_printer[snapshot.printer_id] = snapshot

    if not latest_by_printer:
        return []

    rows = []
    for printer in printers:
        snapshot = latest_by_printer.get(printer.id)
        model_hint = " ".join(
            str(getattr(printer, attr, "") or "")
            for attr in ("shared_name", "name", "model")
        ).upper()
        model_says_color = detect_is_color_from_model(model_hint)
        model_says_mono = (
            "BNW" in model_hint
            or "B/N" in model_hint
            or "B&W" in model_hint
            or "BLACK" in model_hint
            or "MONO" in model_hint
            or bool(re.search(r"\bIM\s*\d{3,5}\b", model_hint))
        ) and not model_says_color
        is_color = False if model_says_mono else (detect_is_color(printer) or model_says_color)
        if not model_says_mono and snapshot and (
            snapshot.color_pages is not None
            or snapshot.copy_color is not None
            or snapshot.print_color is not None
        ):
            is_color = True
        has_counters = bool(
            snapshot
            and (
                snapshot.total_pages is not None
                or snapshot.bw_pages is not None
                or snapshot.color_pages is not None
            )
        )
        color_present = snapshot and is_color and (
            snapshot.color_pages is not None
            or snapshot.copy_color is not None
            or snapshot.print_color is not None
        )
        total_pages = snapshot.total_pages if snapshot else None
        bw_pages = snapshot.bw_pages if snapshot else None
        if snapshot and not is_color and bw_pages is None and total_pages is not None:
            bw_pages = total_pages
        rows.append({
            "id": printer.id,
            "name": getattr(printer, "shared_name", None) or printer.name or printer.model,
            "model": printer.model,
            "serial": printer.serial,
            "ip": printer.ip,
            "is_color": is_color,
            "counter_status": "online" if has_counters else "offline",
            "counter_source": snapshot.source if snapshot else "cache",
            "total_pages": total_pages,
            "copy_bw": snapshot.copy_bw if snapshot else None,
            "copy_color": snapshot.copy_color if color_present else None,
            "print_bw": snapshot.print_bw if snapshot else None,
            "print_color": snapshot.print_color if color_present else None,
            "bw_pages": bw_pages,
            "color_pages": snapshot.color_pages if color_present else None,
            "is_complete": bool(snapshot.is_complete) if snapshot else False,
        })
    return rows


def _refresh_counters_cache(db: Session, force: bool = False) -> tuple[list[dict], float]:
    global COUNTERS_LAST_REFRESH_TS
    now = time.time()
    if (not force and COUNTERS_CACHE["data"] is not None and (now - COUNTERS_CACHE["timestamp"]) < COUNTERS_CACHE_TTL):
        return COUNTERS_CACHE["data"], COUNTERS_CACHE["timestamp"]

    printers = crud.get_printers(db)
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        results = list(executor.map(build_printer_counters, printers))

    captured_at = datetime.now(timezone.utc)
    _upsert_counter_snapshots(db, results, captured_at)

    with COUNTERS_LOCK:
        for row in results:
            printer_id = row.get("id")
            if printer_id is None:
                continue
            prev = COUNTERS_LAST_ROWS.get(printer_id)
            if prev is None or _counter_signature(prev) != _counter_signature(row):
                COUNTERS_LAST_CHANGED_AT[printer_id] = now
            COUNTERS_LAST_ROWS[printer_id] = row
        COUNTERS_LAST_REFRESH_TS = now

    COUNTERS_CACHE["timestamp"] = now
    COUNTERS_CACHE["data"] = results
    return results, now


def _parse_date_yyyy_mm_dd(value: str, field: str) -> datetime:
    try:
        d = datetime.strptime((value or "").strip(), "%Y-%m-%d")
        return datetime(d.year, d.month, d.day, tzinfo=timezone.utc)
    except Exception:
        raise HTTPException(status_code=400, detail=f"{field} inválido. Formato esperado: YYYY-MM-DD")


def _format_report_date(value: datetime | None) -> str | None:
    return value.date().isoformat() if value else None


COUNTER_REPORT_VALUE_FIELDS = (
    "total_pages",
    "bw_pages",
    "color_pages",
    "copy_bw",
    "copy_color",
    "print_bw",
    "print_color",
)


def _snapshot_has_counter_values(snapshot) -> bool:
    if not snapshot:
        return False
    return any(getattr(snapshot, field, None) is not None for field in COUNTER_REPORT_VALUE_FIELDS)


def _first_valid_snapshot(query, start_dt: datetime, end_dt: datetime):
    for snapshot in query.filter(
        models.PrinterCounterSnapshot.captured_at >= start_dt,
        models.PrinterCounterSnapshot.captured_at <= end_dt,
    ).order_by(models.PrinterCounterSnapshot.captured_at.asc()).all():
        if _snapshot_has_counter_values(snapshot):
            return snapshot
    return None


def _last_valid_snapshot(query, end_dt: datetime):
    for snapshot in query.filter(
        models.PrinterCounterSnapshot.captured_at <= end_dt,
    ).order_by(models.PrinterCounterSnapshot.captured_at.desc()).all():
        if _snapshot_has_counter_values(snapshot):
            return snapshot
    return None


def _safe_report_delta(start_value, end_value) -> tuple[int | None, bool]:
    if start_value is None or end_value is None:
        return None, False
    delta = end_value - start_value
    if delta < 0:
        return None, True
    return delta, False


def _build_counters_consumption_report(db: Session, start_date: str, end_date: str) -> dict:
    start_dt = _parse_date_yyyy_mm_dd(start_date, "start_date")
    end_dt = _parse_date_yyyy_mm_dd(end_date, "end_date")
    if end_dt < start_dt:
        raise HTTPException(status_code=400, detail="end_date no puede ser menor que start_date")

    # incluir todo el día final
    end_inclusive = end_dt.replace(hour=23, minute=59, second=59, microsecond=999999)

    printers = crud.get_printers(db)
    report_rows: list[dict] = []

    for printer in printers:
        base_q = db.query(models.PrinterCounterSnapshot).filter(
            models.PrinterCounterSnapshot.printer_id == printer.id,
            models.PrinterCounterSnapshot.granularity == "daily",
        )

        start_snap = _first_valid_snapshot(base_q, start_dt, end_inclusive)
        end_snap = _last_valid_snapshot(base_q, end_inclusive)

        start_total = start_snap.total_pages if start_snap else None
        end_total = end_snap.total_pages if end_snap else None
        start_bw = start_snap.bw_pages if start_snap else None
        end_bw = end_snap.bw_pages if end_snap else None
        start_color = start_snap.color_pages if start_snap else None
        end_color = end_snap.color_pages if end_snap else None

        total_consumed, total_reset = _safe_report_delta(start_total, end_total)
        bw_consumed, bw_reset = _safe_report_delta(start_bw, end_bw)
        color_consumed, color_reset = _safe_report_delta(start_color, end_color)
        if total_consumed is None and bw_consumed is not None and color_consumed is not None:
            total_consumed = bw_consumed + color_consumed

        counter_reset_detected = total_reset or bw_reset or color_reset
        missing_initial = start_snap is None or start_snap.captured_at.date() > start_dt.date()
        missing_final = end_snap is None or end_snap.captured_at.date() < end_dt.date()
        no_comparable_values = (
            total_consumed is None
            and bw_consumed is None
            and color_consumed is None
        )
        notes = []
        if missing_initial:
            notes.append("Sin lectura inicial exacta")
        if missing_final:
            notes.append("Sin lectura final exacta")
        if no_comparable_values:
            notes.append("Sin contadores comparables")
        if counter_reset_detected:
            notes.append("Posible reinicio de contador")

        row = {
            "printer_id": printer.id,
            "name": getattr(printer, "shared_name", None) or printer.name or printer.model,
            "model": printer.model,
            "serial": printer.serial,
            "ip": printer.ip,
            "is_color": bool(getattr(printer, "is_color", False)),
            "start_captured_at": _format_report_date(start_snap.captured_at) if start_snap else None,
            "end_captured_at": _format_report_date(end_snap.captured_at) if end_snap else None,
            "start_total": start_total,
            "end_total": end_total,
            "consumed_total": total_consumed,
            "start_bw": start_bw,
            "end_bw": end_bw,
            "consumed_bw": bw_consumed,
            "start_color": start_color,
            "end_color": end_color,
            "consumed_color": color_consumed,
            "counter_reset_detected": counter_reset_detected,
            "incomplete": missing_initial or missing_final or no_comparable_values or counter_reset_detected,
            "notes": "; ".join(notes),
        }
        report_rows.append(row)

    return {
        "start_date": start_date,
        "end_date": end_date,
        "rows": report_rows,
    }


COUNTERS_REPORT_FIELDS = [
    ("printer_id", "ID"),
    ("name", "Impresora"),
    ("model", "Modelo"),
    ("serial", "Serie"),
    ("ip", "IP"),
    ("is_color", "Tipo color"),
    ("start_captured_at", "Lectura inicial"),
    ("end_captured_at", "Lectura final"),
    ("start_total", "Total inicial"),
    ("end_total", "Total final"),
    ("consumed_total", "Consumo total"),
    ("start_bw", "B/N inicial"),
    ("end_bw", "B/N final"),
    ("consumed_bw", "Consumo B/N"),
    ("start_color", "Color inicial"),
    ("end_color", "Color final"),
    ("consumed_color", "Consumo color"),
    ("counter_reset_detected", "Reinicio contador"),
    ("incomplete", "Incompleto"),
    ("notes", "Notas"),
]


def _format_counters_report_cell(field: str, value):
    if value is None:
        return ""
    if field == "is_color":
        return "Color" if value else "B/N"
    if field in {"counter_reset_detected", "incomplete"}:
        return "Sí" if value else "No"
    return value


@router.get("/counters/consumption-report", dependencies=[Depends(require_tab("counters"))])
def get_counters_consumption_report(
    start_date: str = Query(..., description="YYYY-MM-DD"),
    end_date: str = Query(..., description="YYYY-MM-DD"),
    db: Session = Depends(get_db),
):
    return _build_counters_consumption_report(db, start_date=start_date, end_date=end_date)


@router.get("/counters/consumption-report/export", dependencies=[Depends(require_tab("counters"))])
def export_counters_consumption_report(
    start_date: str = Query(..., description="YYYY-MM-DD"),
    end_date: str = Query(..., description="YYYY-MM-DD"),
    db: Session = Depends(get_db),
):
    report = _build_counters_consumption_report(db, start_date=start_date, end_date=end_date)
    rows = report.get("rows") or []

    output = io.StringIO()
    writer = csv.writer(output)
    fields = [field for field, _ in COUNTERS_REPORT_FIELDS]
    headers = [label for _, label in COUNTERS_REPORT_FIELDS]
    writer.writerow(headers)
    for row in rows:
        writer.writerow([_format_counters_report_cell(field, row.get(field)) for field in fields])

    filename = f"counters_consumption_{start_date}_to_{end_date}.csv"
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/counters/consumption-report/export-xlsx", dependencies=[Depends(require_tab("counters"))])
def export_counters_consumption_report_xlsx(
    start_date: str = Query(..., description="YYYY-MM-DD"),
    end_date: str = Query(..., description="YYYY-MM-DD"),
    db: Session = Depends(get_db),
):
    try:
        from openpyxl import Workbook
    except Exception:
        raise HTTPException(status_code=500, detail="No está disponible la librería openpyxl para exportar XLSX.")

    report = _build_counters_consumption_report(db, start_date=start_date, end_date=end_date)
    rows = report.get("rows") or []

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    fields = [field for field, _ in COUNTERS_REPORT_FIELDS]
    headers = [label for _, label in COUNTERS_REPORT_FIELDS]
    ws.append(headers)

    for row in rows:
        ws.append([_format_counters_report_cell(field, row.get(field)) for field in fields])

    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row_cells in ws.iter_rows(min_row=2):
            for cell in row_cells:
                cell.alignment = Alignment(vertical="top")
        ws.auto_filter.ref = ws.dimensions
    except Exception:
        pass

    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 28)
    ws.freeze_panes = "A2"

    data = io.BytesIO()
    wb.save(data)
    data.seek(0)

    filename = f"counters_consumption_{start_date}_to_{end_date}.xlsx"
    return Response(
        content=data.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def capture_counters_snapshot_now() -> dict:
    db = SessionLocal()
    try:
        rows, ts = _refresh_counters_cache(db, force=True)
        return {
            "captured_count": len(rows or []),
            "timestamp": ts,
        }
    finally:
        db.close()
